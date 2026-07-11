"""
export.py
------------------------------------------------------------------
Bouwt een controleerbaar Excel-exportbestand met de volledige
puntenopbouw van alle deelnemers, voor de beheerder.

Vier tabbladen:
  1. Puntenoverzicht      — matrix: wedstrijd x deelnemer, met som-formules
  2. Voorspellingen       — detail per deelnemer per wedstrijd (audit trail)
  3. Seizoensvoorspelling — na-17 en na-34 voorspelling vs. werkelijkheid
  4. Klassement           — eindtotalen, met formules die naar de andere
                             tabbladen verwijzen (niets hardcoded)

Per-wedstrijd punten komen rechtstreeks uit de database (al berekend door
scoring.py) — die worden niet opnieuw in Excel-formules nagebouwd, want
een tweede, onafhankelijke implementatie van de puntenlogica zou kunnen
gaan afwijken van de app zelf. Alle SOMMEN en TOTALEN in dit bestand zijn
wél live Excel-formules, die verwijzen naar die brongegevens.
------------------------------------------------------------------
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import queries

FONT_NAAM = "Arial"
KOP_FILL = PatternFill(start_color="15161B", end_color="15161B", fill_type="solid")
KOP_FONT = Font(name=FONT_NAAM, bold=True, color="FFFFFF", size=10)
TOTAAL_FILL = PatternFill(start_color="F0F0F3", end_color="F0F0F3", fill_type="solid")
TOTAAL_FONT = Font(name=FONT_NAAM, bold=True, size=10)
CEL_FONT = Font(name=FONT_NAAM, size=10)


def _stijl_kop(cel):
    cel.font = KOP_FONT
    cel.fill = KOP_FILL
    cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _stijl_totaal(cel):
    cel.font = TOTAAL_FONT
    cel.fill = TOTAAL_FILL


def _autofit(ws, breedtes: dict[int, int]):
    for kolom, breedte in breedtes.items():
        ws.column_dimensions[get_column_letter(kolom)].width = breedte


def bouw_export(seizoen: str) -> bytes:
    deelnemers = [
        p for p in queries.get_all_participants()
        if p["status"] == "goedgekeurd"
    ]
    matches = sorted(
        (m for m in queries.get_matches(seizoen) if m["status"] == "afgelopen"),
        key=lambda m: m["kickoff"],
    )

    wb = Workbook()

    # ============================================================
    # Tabblad 2: Voorspellingen (detail/audit — eerst opbouwen, want
    # tabblad 1 verwijst er met formules naar)
    # ============================================================
    ws_detail = wb.active
    ws_detail.title = "Voorspellingen"
    detail_koppen = [
        "Wedstrijd-id", "Datum", "Wedstrijd", "Competitie", "Deelnemer",
        "Voorspelling rust", "Voorspelling eind", "Werkelijke rust", "Werkelijke eind",
        "Joker", "Punten",
    ]
    for kolom, tekst in enumerate(detail_koppen, start=1):
        cel = ws_detail.cell(row=1, column=kolom, value=tekst)
        _stijl_kop(cel)

    rij = 2
    detail_start_rij = rij
    for m in matches:
        for d in deelnemers:
            v = queries.get_my_prediction(m["id"], d["id"])
            waarden = [
                m["id"],
                m["kickoff"].strftime("%d-%m-%Y"),
                f"{m['thuis']} - {m['uit']}",
                m["competitie"],
                d["naam"],
                f"{v['rust_thuis']}-{v['rust_uit']}" if v else "—",
                f"{v['eind_thuis']}-{v['eind_uit']}" if v else "—",
                f"{m['uitslag_rust_thuis']}-{m['uitslag_rust_uit']}",
                f"{m['uitslag_eind_thuis']}-{m['uitslag_eind_uit']}",
                ("Ja" if v and v["joker"] else "Nee") if v else "—",
                v["punten"] if v and v["punten"] is not None else 0,
            ]
            for kolom, waarde in enumerate(waarden, start=1):
                cel = ws_detail.cell(row=rij, column=kolom, value=waarde)
                cel.font = CEL_FONT
            rij += 1
    detail_eind_rij = rij - 1
    _autofit(ws_detail, {1: 10, 2: 11, 3: 26, 4: 16, 5: 18, 6: 16, 7: 16, 8: 16, 9: 16, 10: 8, 11: 9})
    ws_detail.freeze_panes = "A2"

    # ============================================================
    # Tabblad 3: Seizoensvoorspelling
    # ============================================================
    ws_seizoen = wb.create_sheet("Seizoensvoorspelling")
    seizoen_koppen = [
        "Deelnemer",
        "Voorspelling positie na 17", "Voorspelling punten na 17",
        "Werkelijke positie na 17", "Werkelijke punten na 17", "Punten na 17",
        "Voorspelling positie na 34", "Voorspelling punten na 34",
        "Werkelijke positie na 34", "Werkelijke punten na 34", "Punten na 34",
        "Totaal seizoenspunten",
    ]
    for kolom, tekst in enumerate(seizoen_koppen, start=1):
        cel = ws_seizoen.cell(row=1, column=kolom, value=tekst)
        _stijl_kop(cel)

    resultaat = queries.get_season_result(seizoen) or {}
    seizoen_start_rij = 2
    for i, d in enumerate(deelnemers):
        sp = queries.get_season_prediction(d["id"], seizoen)
        r = seizoen_start_rij + i
        ws_seizoen.cell(row=r, column=1, value=d["naam"]).font = CEL_FONT
        ws_seizoen.cell(row=r, column=2, value=sp["na17_positie"] if sp else None).font = CEL_FONT
        ws_seizoen.cell(row=r, column=3, value=sp["na17_punten"] if sp else None).font = CEL_FONT
        ws_seizoen.cell(row=r, column=4, value=resultaat.get("na17_positie")).font = CEL_FONT
        ws_seizoen.cell(row=r, column=5, value=resultaat.get("na17_punten")).font = CEL_FONT
        ws_seizoen.cell(row=r, column=6, value=(sp["punten_na17"] if sp and sp["punten_na17"] is not None else 0)).font = CEL_FONT
        ws_seizoen.cell(row=r, column=7, value=sp["na34_positie"] if sp else None).font = CEL_FONT
        ws_seizoen.cell(row=r, column=8, value=sp["na34_punten"] if sp else None).font = CEL_FONT
        ws_seizoen.cell(row=r, column=9, value=resultaat.get("na34_positie")).font = CEL_FONT
        ws_seizoen.cell(row=r, column=10, value=resultaat.get("na34_punten")).font = CEL_FONT
        ws_seizoen.cell(row=r, column=11, value=(sp["punten_na34"] if sp and sp["punten_na34"] is not None else 0)).font = CEL_FONT
        # Totaal seizoenspunten = live formule, som van de twee checkpoints
        totaal_cel = ws_seizoen.cell(row=r, column=12, value=f"=F{r}+K{r}")
        totaal_cel.font = TOTAAL_FONT
    seizoen_eind_rij = seizoen_start_rij + len(deelnemers) - 1
    _autofit(ws_seizoen, {1: 18, 2: 13, 3: 13, 4: 13, 5: 13, 6: 10, 7: 13, 8: 13, 9: 13, 10: 13, 11: 10, 12: 12})
    ws_seizoen.freeze_panes = "A2"

    # ============================================================
    # Tabblad 1: Puntenoverzicht (matrix, met SUMIFS-formules die naar
    # het Voorspellingen-tabblad verwijzen — dus nooit hardcoded)
    # ============================================================
    ws_overzicht = wb.create_sheet("Puntenoverzicht", 0)
    ws_overzicht.cell(row=1, column=1, value="Datum")
    ws_overzicht.cell(row=1, column=2, value="Wedstrijd")
    ws_overzicht.cell(row=1, column=3, value="Competitie")
    for kolom in (1, 2, 3):
        _stijl_kop(ws_overzicht.cell(row=1, column=kolom))
    for j, d in enumerate(deelnemers):
        kolom = 4 + j
        cel = ws_overzicht.cell(row=1, column=kolom, value=d["naam"])
        _stijl_kop(cel)

    overzicht_start_rij = 2
    for i, m in enumerate(matches):
        r = overzicht_start_rij + i
        ws_overzicht.cell(row=r, column=1, value=m["kickoff"].strftime("%d-%m-%Y")).font = CEL_FONT
        ws_overzicht.cell(row=r, column=2, value=f"{m['thuis']} - {m['uit']}").font = CEL_FONT
        ws_overzicht.cell(row=r, column=3, value=m["competitie"]).font = CEL_FONT
        for j, d in enumerate(deelnemers):
            kolom = 4 + j
            kolomletter = get_column_letter(kolom)
            # SUMIFS: som van Punten in Voorspellingen, waar Wedstrijd-id
            # en Deelnemer overeenkomen met deze rij/kolom.
            formule = (
                f"=SUMIFS(Voorspellingen!$K${detail_start_rij}:$K${detail_eind_rij},"
                f"Voorspellingen!$A${detail_start_rij}:$A${detail_eind_rij},$A{r}&\"\","
                f"Voorspellingen!$E${detail_start_rij}:$E${detail_eind_rij},{kolomletter}$1)"
            )
            # NB: A-kolom hier bevat de datum, niet het wedstrijd-id — daarom
            # gebruiken we hieronder een verborgen wedstrijd-id-kolom in plaats
            # van de datum als matchsleutel (zie kolom na de deelnemers).
        overzicht_eind_rij_placeholder = r
    overzicht_eind_rij = overzicht_start_rij + len(matches) - 1

    # Verborgen kolom met het wedstrijd-id, zodat de SUMIFS-formules hierboven
    # een unieke, betrouwbare sleutel hebben (datum/naam kunnen theoretisch
    # dubbel voorkomen, een id nooit).
    id_kolom = 4 + len(deelnemers)
    ws_overzicht.cell(row=1, column=id_kolom, value="Wedstrijd-id")
    _stijl_kop(ws_overzicht.cell(row=1, column=id_kolom))
    for i, m in enumerate(matches):
        r = overzicht_start_rij + i
        ws_overzicht.cell(row=r, column=id_kolom, value=m["id"]).font = CEL_FONT
    ws_overzicht.column_dimensions[get_column_letter(id_kolom)].hidden = True

    # Nu pas de echte SUMIFS-formules schrijven, verwijzend naar de
    # verborgen id-kolom in plaats van naar de datum.
    idkolomletter = get_column_letter(id_kolom)
    for i, m in enumerate(matches):
        r = overzicht_start_rij + i
        for j, d in enumerate(deelnemers):
            kolom = 4 + j
            kolomletter = get_column_letter(kolom)
            formule = (
                f"=SUMIFS(Voorspellingen!$K${detail_start_rij}:$K${detail_eind_rij},"
                f"Voorspellingen!$A${detail_start_rij}:$A${detail_eind_rij},${idkolomletter}{r},"
                f"Voorspellingen!$E${detail_start_rij}:$E${detail_eind_rij},{kolomletter}$1)"
            )
            ws_overzicht.cell(row=r, column=kolom, value=formule).font = CEL_FONT

    # Totaalrij: som per deelnemer over alle wedstrijden
    totaal_rij = overzicht_eind_rij + 1
    ws_overzicht.cell(row=totaal_rij, column=2, value="Totaal wedstrijdpunten")
    _stijl_totaal(ws_overzicht.cell(row=totaal_rij, column=2))
    for j, d in enumerate(deelnemers):
        kolom = 4 + j
        kolomletter = get_column_letter(kolom)
        cel = ws_overzicht.cell(
            row=totaal_rij, column=kolom,
            value=f"=SUM({kolomletter}{overzicht_start_rij}:{kolomletter}{overzicht_eind_rij})",
        )
        _stijl_totaal(cel)

    breedtes = {1: 12, 2: 26, 3: 16}
    for j in range(len(deelnemers)):
        breedtes[4 + j] = 14
    _autofit(ws_overzicht, breedtes)
    ws_overzicht.freeze_panes = "D2"

    # ============================================================
    # Tabblad 4: Klassement (eindtotalen — alles via formules)
    # ============================================================
    ws_klassement = wb.create_sheet("Klassement")
    klassement_koppen = ["Deelnemer", "Totaal wedstrijdpunten", "Totaal seizoenspunten", "Eindtotaal"]
    for kolom, tekst in enumerate(klassement_koppen, start=1):
        cel = ws_klassement.cell(row=1, column=kolom, value=tekst)
        _stijl_kop(cel)

    for j, d in enumerate(deelnemers):
        r = 2 + j
        overzicht_kolomletter = get_column_letter(4 + j)
        ws_klassement.cell(row=r, column=1, value=d["naam"]).font = CEL_FONT
        # Verwijst naar de totaalrij op het Puntenoverzicht-tabblad
        ws_klassement.cell(
            row=r, column=2,
            value=f"=Puntenoverzicht!{overzicht_kolomletter}{totaal_rij}",
        ).font = CEL_FONT
        # Verwijst naar de bijbehorende rij op het Seizoensvoorspelling-tabblad
        # via INDEX/MATCH op naam (nooit XLOOKUP, zie skill-richtlijnen).
        ws_klassement.cell(
            row=r, column=3,
            value=(
                f"=IFERROR(INDEX(Seizoensvoorspelling!$L${seizoen_start_rij}:$L${seizoen_eind_rij},"
                f"MATCH(A{r},Seizoensvoorspelling!$A${seizoen_start_rij}:$A${seizoen_eind_rij},0)),0)"
            ),
        ).font = CEL_FONT
        totaal_cel = ws_klassement.cell(row=r, column=4, value=f"=B{r}+C{r}")
        totaal_cel.font = TOTAAL_FONT

    _autofit(ws_klassement, {1: 18, 2: 18, 3: 18, 4: 12})
    ws_klassement.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
